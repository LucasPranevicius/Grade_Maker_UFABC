# pip install pandas openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import os

def criar_grade_visual(csv_grade="minha_grade_perfeita.csv", output_filename="Grade_Visual_2026.xlsx"):
    print(f"🎨 Lendo sua grade oficial ({csv_grade}) para montar as tabelas...")
    
    try:
        df_grade = pd.read_csv(csv_grade)
    except FileNotFoundError:
        print(f"❌ Erro: O arquivo '{csv_grade}' não foi encontrado. Rode o 'montar_grade.py' primeiro.")
        return

    wb = Workbook()
    
    wb.remove(wb.active)
    
    dias = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
    horarios = ["08:00-10:00", "10:00-12:00", "14:00-16:00", "16:00-18:00", "19:00-21:00", "21:00-23:00"]
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Azul Escuro
    header_font = Font(color="FFFFFF", bold=True, size=12)
    time_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")   # Azul Cinza
    time_font = Font(color="FFFFFF", bold=True)
    
    cor_semanal = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid") # Azul claro
    cor_q1 = PatternFill(start_color="D1F2EB", end_color="D1F2EB", fill_type="solid")      # Verde claro
    cor_q2 = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")      # Amarelo claro
    
    border_side = Side(border_style="thin", color="BDC3C7")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
    schedule_data = {dia: {h: {"I": None, "II": None, "Semanal": None} for h in horarios} for dia in dias}
    
    for _, row in df_grade.iterrows():
        d = row["Dia"]
        h = row["Horario"]
        q = row["Quinzena"]
        materia_str = f"{row['Nome']}\n({row['Codigo']} - {row['Campus']})"
        
        if d in schedule_data and h in schedule_data[d]:
            schedule_data[d][h][q] = materia_str

    config_abas = [
        ("Quinzena I", "I", cor_q1),
        ("Quinzena II", "II", cor_q2)
    ]
    
    for titulo_aba, chave_quinzena, cor_fundo in config_abas:
        ws = wb.create_sheet(title=titulo_aba)
        
        ws.cell(row=1, column=1, value="Horário / Dia").fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = border
        ws.column_dimensions["A"].width = 16
        
        for col_idx, dia in enumerate(dias, start=2):
            cell = ws.cell(row=1, column=col_idx, value=dia)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 30 
            
        grid_map = {}
        for row_idx, horario in enumerate(horarios, start=2):
            cell = ws.cell(row=row_idx, column=1, value=horario)
            cell.fill = time_fill
            cell.font = time_font
            cell.alignment = center_align
            cell.border = border
            ws.row_dimensions[row_idx].height = 65 
            grid_map[horario] = row_idx
            
            for col_idx in range(2, len(dias) + 2):
                ws.cell(row=row_idx, column=col_idx).border = border
                
        for dia_idx, dia in enumerate(dias, start=2):
            for horario, row_idx in grid_map.items():
                cell = ws.cell(row=row_idx, column=dia_idx)
                cell.alignment = center_align
                cell.border = border
                
                slots = schedule_data[dia][horario]
                
                if slots["Semanal"]:
                    cell.value = slots["Semanal"]
                    cell.fill = cor_semanal
                
                elif slots[chave_quinzena]:
                    cell.value = f"[{titulo_aba}]\n{slots[chave_quinzena]}"
                    cell.fill = cor_fundo

    wb.save(output_filename)
    print(f"\n✅ Arquivo '{output_filename}' gerado com duas abas separadas (Quinzena I e Quinzena II)!")

if __name__ == "__main__":
    criar_grade_visual()